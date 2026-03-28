"""
Generate sample purchase register Excel file for testing.
Run: python generate_test_data.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def generate_sample_purchase_register():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Register"

    # Headers (Tally-style)
    headers = [
        "Voucher Date", "Voucher No", "Party Name", "Party GSTIN",
        "Taxable Value", "IGST", "CGST", "SGST"
    ]

    # Style headers
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data — matches some GSTR-2B entries, some missing, some with value differences
    data = [
        # Matches with GSTR-2B
        ["05-03-2026", "INV-2026-0834", "Reliance Steel Ltd.", "27AAACR5678B1Z3", 450000, 81000, 0, 0],
        ["10-03-2026", "INV-2026-0901", "Reliance Steel Ltd.", "27AAACR5678B1Z3", 320000, 57600, 0, 0],

        # Value mismatch — PR has 280000, 2B has 250000
        ["12-03-2026", "TC/23-24/1567", "Tata Chemicals Pvt Ltd", "24AAACT1234C1Z1", 280000, 50400, 0, 0],

        # Missing in 2B — these invoices won't be found in GSTR-2B
        ["08-03-2026", "JSW-029384", "JSW Paints Dealers", "29AABCJ9012D1Z7", 195000, 35100, 0, 0],
        ["15-03-2026", "BEC-2026-445", "Bharat Electricals Corp", "27AABCB3456E1Z9", 167000, 30060, 0, 0],

        # Value mismatch — PR has 134000, 2B has 120000
        ["03-03-2026", "AMB/24/09823", "Ambuja Cements Ltd", "24AAACA7890F1Z2", 134000, 24120, 0, 0],

        # Missing in 2B — different invoice number from what's in 2B
        ["18-03-2026", "SI-MAR-2026-112", "Supreme Industries", "27AABCS4567G1Z4", 112000, 20160, 0, 0],

        # Value mismatch — PR has 98000, 2B has 88000
        ["10-03-2026", "LT-WO-56789", "Larsen & Toubro", "27AAACL8901H1Z6", 98000, 17640, 0, 0],

        # Missing in 2B
        ["20-03-2026", "AP/26/MH/4421", "Asian Paints Ltd", "27AAACA2345I1Z8", 76000, 13680, 0, 0],

        # Value mismatch — PR has 45000, 2B has 40000
        ["22-03-2026", "HAV/MAR/8821", "Havells India Ltd", "07AABCH1234K1Z3", 45000, 8100, 0, 0],

        # Additional matching entries
        ["01-03-2026", "SI-MAR-2026-100", "Supreme Industries", "27AABCS4567G1Z4", 185000, 33300, 0, 0],

        # More entries that match perfectly
        ["25-03-2026", "PQR-2026-001", "PQR Traders", "27AABCP1234L1Z5", 88000, 15840, 0, 0],
        ["26-03-2026", "STU-2026-045", "STU Enterprises", "24AABCS5678M1Z7", 56000, 10080, 0, 0],
        ["27-03-2026", "VWX-2026-112", "VWX Industries", "29AABCV9012N1Z9", 42000, 7560, 0, 0],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx >= 5:  # Numeric columns
                cell.number_format = '#,##0'

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

    # Save
    filepath = os.path.join(os.path.dirname(__file__), "sample_purchase_register.xlsx")
    wb.save(filepath)
    print(f"Generated: {filepath}")
    print(f"Total invoices: {len(data)}")
    print(f"Expected mismatches: ~7 (4 missing in 2B, 3 value mismatches)")


if __name__ == "__main__":
    generate_sample_purchase_register()
